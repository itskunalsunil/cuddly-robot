import pandas as pd
import webbrowser
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys

# Path to Excel file
excel_file = "C:/Users/Kunal.Sunil/OneDrive - Investcorp-Tages Ltd/Desktop/New folder/All Contacts - June 2025.xlsx"

df = pd.read_excel(excel_file, engine='openpyxl')

if 'Search Status' not in df.columns:
    df['Search Status'] = ""

df_subset = df.iloc[3769:6666]  
next_index = df_subset[df_subset['Search Status'].isna()].index.min()

if pd.isna(next_index):
    sys.exit("All rows have been processed or no rows with missing 'Search Status' found.")
else:
    print(f"Next row to process: {next_index}")

# If you want to start at row 3453 in xl sheet, make lower bound 3451. 
# To end at 3473, make upper bound 3472
# The index it gives you is sufficient, just copy it over to lower bound
for index, row in df.iloc[next_index:6666].iterrows():
    first_name = row['First Name']
    last_name = row['Last Name']
    company = row['Account Name']

    if pd.isna(first_name) or pd.isna(last_name) or pd.isna(company):
        print(f"Skipping row {index} due to missing data.")
        continue

    name = f"{first_name} {last_name}"
    query = f"{first_name} {last_name} {company}"
    search_url = f"https://www.linkedin.com/search/results/people/?keywords={query.replace(' ', '%20')}"

    webbrowser.open(search_url)
    print(f"\n🔍 Opened search for: {name} at {company}")

    status = input("Enter status (Not Found, Changed Company, OK, Duplicate) or type 'exit' to stop: ").strip()

    if status.lower() == "exit":
        print("Exiting early by user request.")
        print(f"\nYou got to line: {index}")
        break

    if status == "Not Found" or status == "not found" or status == "Not found":
        google_url = f"https://www.google.com/search?q={query.replace(' ', '+')}"
        webbrowser.open(google_url)
        print("Opened google search to check further")
        status = input("Anything new? (Not Found, Changed Company, OK, Duplicate): ").strip()

    df.at[index, 'Search Status'] = status

    df.to_excel(excel_file, index=False, engine='openpyxl')


wb = load_workbook(excel_file)
ws = wb.active

red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
green_fill = PatternFill(start_color="39FF14", end_color="39FF14", fill_type="solid")
clear_fill = PatternFill(fill_type=None)

status_col = None
for col in ws.iter_cols(1, ws.max_column):
    if col[0].value == 'Search Status':
        status_col = col[0].column
        break

if status_col:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=status_col)
        if cell.value == "Not Found" or cell.value == "Not found" or cell.value == "not found":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = red_fill
        elif cell.value == "Changed Company" or cell.value == "changed company":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill
        elif cell.value == "OK" or cell.value == "ok":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = clear_fill
        elif cell.value == "Duplicate" or cell.value == "duplicate":
            for col in range(1,ws.max_column + 1):
                ws.cell(row=row, column=col).fill = green_fill

wb.save(excel_file)

print("✅ Excel file updated with statuses")
